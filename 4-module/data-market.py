import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string

def sales_report(file_name):
    # Extraemos el mes y extensión del nombre del archivo
    mes_extension = file_name.split('_')[1]

    # Input Sales_mes.xlsx / Output report_sales_mes.xlsx
    data = pd.read_excel(f'../datasets/{file_name}')
    data_filt = data[['Gender','Product line','Total']]
    data_pivot = data_filt.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    data_pivot.to_excel(f'./outputs/sales_{mes_extension}', startrow=4, sheet_name='Report')

    # Escribiendo datos en excel con openpyxl
    wb = load_workbook(f'./outputs/sales_{mes_extension}')
    sheet1 = wb['Report']

    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Insertamos un gráfico en el archivo Excel
    barchart = BarChart()
    excel_data = Reference(sheet1, min_col=min_col+1, max_col=max_col, min_row=min_row,  max_row=max_row)
    categorias = Reference(sheet1, min_col=min_col,  max_col=min_col, min_row=min_row+1, max_row=max_row)

    barchart.add_data(excel_data, titles_from_data=True)
    barchart.set_categories(categorias)

    sheet1.add_chart(barchart, 'B12')
    barchart.title = 'Ventas'
    barchart.style = 2

    # Extraemos el mes de la variable mes_extensión
    mes = mes_extension.split('.')[0]

    # Añadimos titulo y subtitulo
    sheet1['A1'].font = Font('Arial', bold=True, size=20)
    sheet1['A2'].font = Font('Arial', bold=True, size=12)
    sheet1['A1'] = 'Reporte en excel desde Python'
    sheet1['A2'] = f'Año 2021 {mes}'

    # Generamos el Reporte
    sheet1['B8'] = '=SUM(B6:B7)'
    sheet1['B8'].style ='Currency'
    sheet1[f'A{max_row+1}'] ='Total'

    # Añadimos totales en cada columna
    abecedario = list(string.ascii_uppercase)
    for i in abecedario[1:max_col]:
        sheet1[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1} : {i}{max_row})'
        sheet1[f'{i}{max_row+1}'].style ='Currency'   
        
    # salvamos el archivo con los cambios realizados
    wb.save(f'./outputs/sales_{mes_extension}')
    print('done')

# Format = Input Sales_mes.xlsx / Output sales_mes_report.xlsx
sales_report('market_05.xlsx')

